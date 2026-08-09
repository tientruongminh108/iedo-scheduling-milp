"""
Data structures + Parser for problem instance.
"""
from __future__ import annotations

import datetime
import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import List

from openpyxl import load_workbook     # type: ignore


@dataclass
class Job:
    """A single job (one row) within a problem instance."""
    job_id: int
    process_types: List[str]        # e.g. ["Boiling", "Baking", "Smoking"]
    processing_times: List[float]   # hours, same length as process_types
    splitting_timing: int           # 0 = cannot be split; k = may split after process k
    due_time: float                 # hours after opening

    @property
    def id(self) -> int:
        """Alias for job_id to support concise variable indexing."""        
        return self.job_id
    
    def get_piece_time(self, piece: int) -> float:
        """
        Calculate total processing time for Piece 1 or Piece 2.

        - If job cannot split: Piece 1 gets total processing time, Piece 2 gets 0.0.
        - If split after process k: Piece 1 gets processes 0..k-1, Piece 2 gets processes k..end.
        """
        if not self.can_split:
            return self.total_processing_time if piece == 1 else 0.0

        k = self.splitting_timing
        if piece == 1:
            return sum(self.processing_times[:k])
        elif piece == 2:
            return sum(self.processing_times[k:])
        return 0.0

    @property
    def num_process(self) -> int:
        return len(self.process_types)

    @property
    def total_processing_time(self) -> float:
        return sum(self.processing_times)

    @property
    def can_split(self) -> bool:
        return self.splitting_timing > 0

    def to_dict(self) -> dict:
        return {
            "job_id": self.job_id,
            "process_types": self.process_types,
            "processing_times": self.processing_times,
            "splitting_timing": self.splitting_timing,
            "due_time": self.due_time,
        }

    @staticmethod
    def from_dict(d: dict) -> "Job":
        return Job(
            job_id=d["job_id"],
            process_types=list(d["process_types"]),
            processing_times=list(d["processing_times"]),
            splitting_timing=d["splitting_timing"],
            due_time=d["due_time"],
        )


@dataclass
class Machine:
    """A processing machine with specific capabilities."""
    id: int
    capabilities: List[str]  # e.g., ["Boiling"], or ["All"]

    def can_process(self, job: Job) -> bool:
        """Check if this machine can perform all processes required by the job."""
        if "All" in self.capabilities:
            return True
        # If any of the job's processes is not in this machine's capabilities, it can't process it
        for p_type in job.process_types:
            if p_type not in self.capabilities:
                return False
        return True


@dataclass
class Instance:
    """One full problem instance."""
    name: str     # e.g. "Instance 1"
    jobs: List[Job] = field(default_factory=list)
    machines: List[Machine] = field(default_factory=list)

    def __post_init__(self):
        # Automatically populate the default 5 machines for Case 1 if empty
        if not self.machines:
            self.machines = self.default_machines()

    @staticmethod
    def default_machines() -> List[Machine]:
        """Return the default machine configuration for Case 1."""
        return [
            Machine(id=1, capabilities=["Boiling"]),
            Machine(id=2, capabilities=["All"]),
            Machine(id=3, capabilities=["All"]),
            Machine(id=4, capabilities=["All"]),
            Machine(id=5, capabilities=["All"]),
        ]

    @property
    def num_jobs(self) -> int:
        return len(self.jobs)

    def to_dict(self) -> dict:
        return {"name": self.name, "jobs": [j.to_dict() for j in self.jobs], "machines": [{"id": m.id, "capabilities": m.capabilities} for m in self.machines]}

    @staticmethod
    def from_dict(d: dict) -> "Instance":
        machines = []
        for m in d.get("machines", []):
            machines.append(Machine(id=m["id"], capabilities=m["capabilities"]))
        return Instance(name=d["name"], jobs=[Job.from_dict(j) for j in d["jobs"]], machines=machines)



def _time_to_hours(value) -> float:
    """Convert an Excel time-of-day cell to hours after 7:30 AM (opening time).
    
    Handles:
    - datetime.time objects
    - datetime.datetime objects  
    - Excel serial date/time (float): fractional part = time of day, integer part = date
    - Raw hours (float/int)
    """
    abs_hours = 0.0
    if isinstance(value, datetime.time):
        abs_hours = value.hour + value.minute / 60 + value.second / 3600
    elif isinstance(value, datetime.datetime):
        t = value.time()
        abs_hours = t.hour + t.minute / 60 + t.second / 3600
    elif isinstance(value, (int, float)):
        # Excel serial date/time: integer part = days since 1900-01-00, fractional part = time of day
        # If value >= 1, it's a date+time serial; extract fractional part for time of day
        # If value < 1, it's a pure time serial (fraction of 24h)
        if value >= 1:
            # Date+time: extract fractional part (time of day)
            frac = value - int(value)
            abs_hours = frac * 24
        else:
            # Pure time serial
            abs_hours = float(value) * 24
    elif isinstance(value, str):
        # Parse str HH:MM or HH:MM:SS
        parts = [float(p) for p in value.strip().split(":")]
        abs_hours = parts[0] + parts[1]/60 + (parts[2]/3600 if len(parts) > 2 else 0)
    else:
        raise TypeError(f"Unrecognized due-time cell value: {value!r} ({type(value)})")
    
    # 0 = 7:30 AM
    relative_hours = abs_hours - 7.5 
    
    return relative_hours if relative_hours >= 0 else relative_hours + 24.0


def _parse_sheet(ws) -> Instance:
    rows = list(ws.iter_rows(values_only=True))
    header_1, header_2 = rows[0], rows[1]
    data_rows = rows[2:]
    
    split_col = next(i for i, v in enumerate(header_1) if v == "Splitting Timing")
    due_col = next(i for i, v in enumerate(header_1) if v == "Due Time")

    process_type_cols = [i for i in range(1, split_col) if header_2[i] is not None]
    n_proc = len(process_type_cols)     # number of processes
    proc_time_cols = list(range(split_col + 1, split_col + 1 + n_proc))

    jobs: List[Job] = []
    for row in data_rows:
        if row[0] is None:
            continue    # skip stray blank rows, if any

        types = []
        times = []
        for tcol, pcol in zip(process_type_cols, proc_time_cols):
            ptype = row[tcol]
            ptime = row[pcol]
            if ptype is None:
                continue    # this job has fewer processes than the max

            types.append(str(ptype))
            times.append(float(ptime))
        
        jobs.append(
            Job(
                job_id=int(row[0]),
                process_types=types,
                processing_times=times,
                splitting_timing=int(row[split_col]),
                due_time=_time_to_hours(row[due_col])
            )
        )

    return Instance(name=ws.title, jobs=jobs)


def parse_workbook(path: str | Path) -> List[Instance]:
    """Read every 'Instance *' sheet in the workbook into Instance objects."""
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_names = [s for s in wb.sheetnames if s.lower().startswith("instance")]
    if not sheet_names:
        raise ValueError(
            f"No sheet named like 'Instance N' found in {path}. "
            f"Sheets present: {wb.sheetnames}"
        )
    return [_parse_sheet(wb[name]) for name in sheet_names]


def cache_instances(instances: List[Instance], out_dir: str | Path) -> List[Path]:
    """Write one JSON file per instance to out_dir. Returns the written paths."""
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    written = []
    for i, inst in enumerate(instances, start=1):
        path = out_dir / f"instance_{i:02d}.json"
        path.write_text(json.dumps(inst.to_dict(), indent=2), encoding="utf-8")
        written.append(path)
    return written


def load_cached_instances(processed_dir: str | Path) -> List[Instance]:
    """Read back instance_*.json files, e.g. to skip re-parsing Excel."""
    processed_dir = Path(processed_dir)
    files = sorted(processed_dir.glob("instance_*.json"))
    if not files:
        raise FileNotFoundError(f"No cached instance_*.json files found in {processed_dir}")
    return [Instance.from_dict(json.loads(f.read_text(encoding="utf-8"))) for f in files]